import json
import sys
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl missing"}))
    sys.exit(1)


import difflib

def find_column(header, keywords):
    # Try exact and partial match first
    for idx, title in enumerate(header):
        text = (title or "").strip().lower()
        for keyword in keywords:
            if keyword in text or text in keyword:
                return idx
    # Fuzzy match if not found
    for idx, title in enumerate(header):
        text = (title or "").strip().lower()
        close = difflib.get_close_matches(text, keywords, n=1, cutoff=0.6)
        if close:
            return idx
    return None


def main():
    if len(sys.argv) < 2:
        print("[]")
        return

    path = sys.argv[1]
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[]")
        return

    header = [(cell or "").strip() if isinstance(cell, str) else str(cell or "").strip() for cell in rows[0]]
    lowered = [h.lower() for h in header]

    idx_college = find_column(lowered, ["college", "institution", "institute"])
    idx_branch = find_column(lowered, ["branch", "course", "program"])
    idx_category = find_column(lowered, ["category", "quota"])
    idx_cutoff = find_column(lowered, ["cutoff", "closing", "rank"])

    if None in (idx_college, idx_branch, idx_category, idx_cutoff):
        print("[]")
        return

    data = []
    for row in rows[1:]:
        college = (row[idx_college] or "").strip() if row[idx_college] else ""
        branch = (row[idx_branch] or "").strip() if row[idx_branch] else ""
        category = (row[idx_category] or "").strip() if row[idx_category] else ""
        cutoff_value = row[idx_cutoff]
        if isinstance(cutoff_value, (int, float)):
            cutoff = int(cutoff_value)
        else:
            cutoff = int(re.sub(r"[^0-9]", "", str(cutoff_value))) if cutoff_value else 0

        data.append({
            "college": college,
            "branch": branch,
            "category": category,
            "cutoff": cutoff
        })

    print(json.dumps(data, ensure_ascii=False))


if __name__ == "__main__":
    main()

